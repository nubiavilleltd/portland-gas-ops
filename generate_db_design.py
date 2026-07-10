"""
Portland Gas DB Design — Full Schema Generator
Run: python generate_db_design.py
Output: Portland_Gas_DB_Design.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "purple":      "7234BD",
    "purple_lt":   "F3EEFF",
    "gold":        "FFBC00",
    "gold_lt":     "FFFBEA",
    "dark":        "1C043B",
    "green":       "166534",
    "green_lt":    "F0FDF4",
    "blue":        "1E40AF",
    "blue_lt":     "EFF6FF",
    "red":         "991B1B",
    "red_lt":      "FEF2F2",
    "amber":       "92400E",
    "amber_lt":    "FFFBEB",
    "teal":        "0F766E",
    "teal_lt":     "F0FDFA",
    "slate":       "334155",
    "slate_lt":    "F8FAFC",
    "header_row":  "1C043B",
    "col_header":  "2D1057",
    "white":       "FFFFFF",
    "gray_border": "E2E8F0",
    "pk":          "FFBC00",
    "fk":          "C4B5FD",
    "null_col":    "94A3B8",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="1C043B", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def border():
    s = Side(style="thin", color=C["gray_border"])
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Groups: (sheet_tab_name, display_title, accent_hex, tables_list) ──────────
# Each table = (table_name, description, [(col, type, nullable, pk/fk/notes)])
# Set hidden=True on any group to hide its sheet and exclude it from the Index.

GROUPS = [

    # ── 0. INDEX ──────────────────────────────────────────────────────────────
    None,   # placeholder, handled separately

    # ── 1. IDENTITY  (sheet hidden by default) ────────────────────────────────
    {
        "sheet_name":    "Identity",
        "display_title": "Identity & Organisation",
        "accent":        C["purple"],
        "hidden":        True,
        "tables": [
            (
                "users",
                "Authentication records only — one row per login account",
                [
                    ("id",             "SERIAL",      "NO",  "PK"),
                    ("first_name",     "VARCHAR(80)", "NO",  ""),
                    ("last_name",      "VARCHAR(80)", "NO",  ""),
                    ("email",          "VARCHAR(150)","NO",  "UNIQUE"),
                    ("password_hash",  "TEXT",        "NO",  "bcrypt"),
                    ("is_active",      "BOOLEAN",     "NO",  "default true"),
                    ("last_login",     "TIMESTAMPTZ", "YES", ""),
                    ("created_at",     "TIMESTAMPTZ", "NO",  "default now()"),
                    ("updated_at",     "TIMESTAMPTZ", "NO",  "default now()"),
                ],
            ),
            (
                "employees",
                "Business/HR identity — bridges to users; all workflow FKs point here",
                [
                    ("id",              "SERIAL",      "NO",  "PK"),
                    ("user_id",         "INT",         "NO",  "FK → users.id · UNIQUE"),
                    ("employee_no",     "VARCHAR(20)", "NO",  "UNIQUE  e.g. PGL-001"),
                    ("department",      "VARCHAR(100)","NO",  ""),
                    ("job_title",       "VARCHAR(100)","NO",  ""),
                    ("role",            "VARCHAR(80)", "NO",  "e.g. finance_manager, md, line_manager"),
                    ("line_manager_id", "INT",         "YES", "FK → employees.id (self-ref)"),
                    ("employment_type", "VARCHAR(30)", "NO",  "full_time | contract | intern"),
                    ("phone",           "VARCHAR(20)", "YES", ""),
                    ("hire_date",       "DATE",        "NO",  ""),
                    ("is_active",       "BOOLEAN",     "NO",  "default true"),
                    ("created_at",      "TIMESTAMPTZ", "NO",  "default now()"),
                    ("updated_at",      "TIMESTAMPTZ", "NO",  "default now()"),
                ],
            ),
        ],
    },

    # ── 2. DOCUMENTS ─────────────────────────────────────────────────────────
    {
        "sheet_name":    "Documents",
        "display_title": "Document Library",
        "accent":        C["teal"],
        "hidden":        False,
        "tables": [
            (
                "documents",
                "Unified file & folder store — self-referencing parent_id for unlimited nesting. type='folder' has no file_path.",
                [
                    ("id",           "SERIAL",       "NO",  "PK"),
                    ("parent_id",    "INT",          "YES", "FK → documents.id  (null = root)"),
                    ("type",         "VARCHAR(10)",  "NO",  "file | folder"),
                    ("name",         "VARCHAR(255)", "NO",  "display name / filename"),
                    ("category",     "VARCHAR(80)",  "YES", "procurement | asset | hr | fleet | general …"),
                    ("file_path",    "TEXT",         "YES", "S3 key or Cloudinary URL — null for folders"),
                    ("file_size",    "BIGINT",       "YES", "bytes — null for folders"),
                    ("mime_type",    "VARCHAR(100)", "YES", "null for folders"),
                    ("uploaded_by",  "INT",          "YES", "FK → employees.id"),
                    ("created_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
        ],
    },

    # ── 3. APPROVAL WORKFLOW ENGINE ───────────────────────────────────────────
    {
        "sheet_name":    "Approval",
        "display_title": "Approval Workflow Engine",
        "accent":        C["blue"],
        "hidden":        False,
        "tables": [
            (
                "approval_workflows",
                "Named workflow templates — one per process type",
                [
                    ("id",          "SERIAL",       "NO",  "PK"),
                    ("name",        "VARCHAR(100)", "NO",  "e.g. Procurement Request, Asset Request"),
                    ("description", "TEXT",         "YES", ""),
                    ("is_active",   "BOOLEAN",      "NO",  "default true"),
                    ("created_at",  "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "workflow_steps",
                "Ordered steps within a workflow template. assignee_type drives how the approver is resolved at runtime.",
                [
                    ("id",              "SERIAL",      "NO",  "PK"),
                    ("workflow_id",     "INT",         "NO",  "FK → approval_workflows.id"),
                    ("step_number",     "INT",         "NO",  "1-based ordering"),
                    ("step_name",       "VARCHAR(100)","NO",  "e.g. Line Manager Review"),
                    ("assignee_type",   "VARCHAR(40)", "NO",  "role | specific | requester_pick | requester_line_manager | requester_hod | requester_skip_level"),
                    ("role",            "VARCHAR(80)", "YES", "resolved at runtime when assignee_type=role  e.g. finance_manager"),
                    ("employee_id",     "INT",         "YES", "FK → employees.id  (when assignee_type=specific)"),
                    ("group_id",        "INT",         "YES", "FK → approver_groups.id  (when assignee_type=requester_pick)"),
                    ("can_approve",     "BOOLEAN",     "NO",  "default true — show Approve button for this step"),
                    ("can_reject",      "BOOLEAN",     "NO",  "default true — show Reject button for this step"),
                    ("can_return",      "BOOLEAN",     "NO",  "default false — show Return to Initiator button"),
                    ("can_escalate",    "BOOLEAN",     "NO",  "default false — show Escalate button"),
                    ("created_at",      "TIMESTAMPTZ", "NO",  "default now()"),
                ],
            ),
            (
                "approval_requests",
                "Live tracker — one row per request entering a workflow. current_step_number advances on each approval.",
                [
                    ("id",                  "SERIAL",      "NO",  "PK"),
                    ("workflow_id",         "INT",         "NO",  "FK → approval_workflows.id"),
                    ("request_type",        "VARCHAR(50)", "NO",  "procurement | asset"),
                    ("request_id",          "INT",         "NO",  "PK of the originating request row"),
                    ("submitted_by",        "INT",         "NO",  "FK → employees.id"),
                    ("current_step_number", "INT",         "NO",  "which step is awaiting action"),
                    ("overall_status",      "VARCHAR(20)", "NO",  "pending | approved | rejected | returned"),
                    ("attempt_number",      "INT",         "NO",  "default 1 — increments each time request is resubmitted after return"),
                    ("created_at",          "TIMESTAMPTZ", "NO",  "default now()"),
                    ("updated_at",          "TIMESTAMPTZ", "NO",  "default now()"),
                ],
            ),
            (
                "approval_step_assignments",
                "Stores the resolved approver for each step of each request. Created when a step becomes active.",
                [
                    ("id",                  "SERIAL",     "NO",  "PK"),
                    ("approval_request_id", "INT",        "NO",  "FK → approval_requests.id"),
                    ("step_number",         "INT",        "NO",  ""),
                    ("assigned_to",         "INT",        "NO",  "FK → employees.id — the actual person"),
                    ("assigned_at",         "TIMESTAMPTZ","NO",  "when this assignment was created"),
                ],
            ),
            (
                "approval_history",
                "Immutable step-by-step log of approver actions — one row per approval action. Never updated, only inserted.",
                [
                    ("id",                  "SERIAL",      "NO",  "PK"),
                    ("approval_request_id", "INT",         "NO",  "FK → approval_requests.id"),
                    ("step_number",         "INT",         "NO",  ""),
                    ("actor_id",            "INT",         "NO",  "FK → employees.id"),
                    ("action",              "VARCHAR(20)", "NO",  "approved | rejected | returned | escalated"),
                    ("comment",             "TEXT",        "YES", "reason or note from approver"),
                    ("acted_at",            "TIMESTAMPTZ", "NO",  "default now()"),
                ],
            ),
            (
                "workflow_audit_trail",
                "Standalone audit trail written by workflowEngine.audit(). Captures every actor event across all workflow requests — both requester actions (submitted, recalled, returned) and approver actions (approved, rejected, escalated). Called independently from the engine so any service can write to it.",
                [
                    ("id",              "SERIAL",       "NO",  "PK"),
                    ("workflow_id",     "INT",          "NO",  "FK → approval_workflows.id"),
                    ("request_id",      "INT",          "NO",  "PK of the originating request row"),
                    ("request_type",    "VARCHAR(50)",  "NO",  "procurement | asset"),
                    ("actor_id",        "INT",          "NO",  "FK → employees.id — person who performed the action"),
                    ("actor_role",      "VARCHAR(80)",  "YES", "role in this request e.g. requester | line_manager | hod | finance_manager | md"),
                    ("step_number",     "INT",          "YES", "null for requester-initiated actions (submitted, recalled)"),
                    ("action",          "VARCHAR(40)",  "NO",  "submitted | approved | rejected | returned | escalated | recalled"),
                    ("comment",         "TEXT",         "YES", "approver note or requester comment"),
                    ("acted_at",        "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "approver_groups",
                "Named lists of people for requester_pick steps — not role-based, manually curated by admin",
                [
                    ("id",          "SERIAL",       "NO",  "PK"),
                    ("name",        "VARCHAR(100)", "NO",  "e.g. SODA Approvers, Executive Committee"),
                    ("description", "TEXT",         "YES", ""),
                    ("is_active",   "BOOLEAN",      "NO",  "default true"),
                    ("created_by",  "INT",          "NO",  "FK → employees.id"),
                    ("created_at",  "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",  "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "approver_group_members",
                "Members belonging to an approver group — many employees can belong to many groups",
                [
                    ("id",          "SERIAL",      "NO",  "PK"),
                    ("group_id",    "INT",         "NO",  "FK → approver_groups.id"),
                    ("employee_id", "INT",         "NO",  "FK → employees.id"),
                    ("added_at",    "TIMESTAMPTZ", "NO",  "default now()"),
                ],
            ),
            (
                "workflow_assignments",
                "Maps each process/request_type to its active workflow — tells the engine which workflow to trigger",
                [
                    ("id",           "SERIAL",      "NO",  "PK"),
                    ("request_type", "VARCHAR(50)", "NO",  "UNIQUE  procurement | asset"),
                    ("workflow_id",  "INT",         "NO",  "FK → approval_workflows.id"),
                    ("is_active",    "BOOLEAN",     "NO",  "default true"),
                    ("updated_by",   "INT",         "NO",  "FK → employees.id"),
                    ("updated_at",   "TIMESTAMPTZ", "NO",  "default now()"),
                ],
            ),
        ],
    },

    # ── 4. ALL REQUESTS ───────────────────────────────────────────────────────
    {
        "sheet_name":    "All Requests",
        "display_title": "All Requests",
        "accent":        C["slate"],
        "hidden":        False,
        "tables": [
            (
                "all_requests",
                "Unified registry of every request across all processes — powers the All Requests dashboard. Updated on each status change.",
                [
                    ("id",                  "SERIAL",       "NO",  "PK"),
                    ("reference",           "VARCHAR(20)",  "NO",  "UNIQUE  e.g. PR-2026-001, AR-2026-001  — human-readable identifier shown to users"),
                    ("request_type",        "VARCHAR(50)",  "NO",  "procurement | asset"),
                    ("request_id",          "INT",          "NO",  "PK of the originating request row"),
                    ("title",               "VARCHAR(200)", "NO",  ""),
                    ("raised_by",           "INT",          "NO",  "FK → employees.id"),
                    ("department",          "VARCHAR(100)", "NO",  ""),
                    ("status",              "VARCHAR(20)",  "NO",  "draft | pending | approved | rejected | returned"),
                    ("approval_request_id", "INT",          "YES", "FK → approval_requests.id  (null until submitted)"),
                    ("created_at",          "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",          "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
        ],
    },

    # ── 5. PROCUREMENT ────────────────────────────────────────────────────────
    {
        "sheet_name":    "Procurement",
        "display_title": "Procurement",
        "accent":        C["green"],
        "hidden":        False,
        "tables": [
            (
                "vendors",
                "Unified vendor registry — permanent (approved company suppliers) and temporary (ad-hoc one-time vendors entered at request time). Filter by vendor_type to separate them. procurement_request_id is null for permanent vendors.",
                [
                    ("id",                     "SERIAL",       "NO",  "PK"),
                    ("vendor_type",            "VARCHAR(15)",  "NO",  "permanent | temporary"),
                    ("name",                   "VARCHAR(200)", "NO",  ""),
                    ("contact_name",           "VARCHAR(150)", "YES", ""),
                    ("email",                  "VARCHAR(150)", "YES", ""),
                    ("phone",                  "VARCHAR(20)",  "YES", ""),
                    ("address",                "TEXT",         "YES", ""),
                    ("bank_name",              "VARCHAR(100)", "YES", ""),
                    ("bank_account_no",        "VARCHAR(30)",  "YES", ""),
                    ("bank_account_name",      "VARCHAR(150)", "YES", ""),
                    ("reason",                 "TEXT",         "YES", "why a registered vendor was not used — populated for temporary vendors only"),
                    ("procurement_request_id", "INT",          "YES", "FK → procurement_requests.id  (null for permanent vendors)"),
                    ("is_active",              "BOOLEAN",      "NO",  "default true — set false to deactivate a permanent vendor"),
                    ("added_by",               "INT",          "NO",  "FK → employees.id"),
                    ("created_at",             "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "procurement_requests",
                "Purchase/procurement request raised by any employee",
                [
                    ("id",               "SERIAL",       "NO",  "PK"),
                    ("reference",        "VARCHAR(20)",  "NO",  "UNIQUE  e.g. PR-2026-001  — human-readable identifier shown to users"),
                    ("raised_by",        "INT",          "NO",  "FK → employees.id"),
                    ("department",       "VARCHAR(100)", "NO",  ""),
                    ("title",            "VARCHAR(200)", "NO",  ""),
                    ("description",      "TEXT",         "YES", ""),
                    ("estimated_amount", "NUMERIC(15,2)","YES", ""),
                    ("currency",         "VARCHAR(5)",   "NO",  "default NGN"),
                    ("vendor_id",        "INT",          "YES", "FK → vendors.id  (permanent or temporary — set after vendor selection)"),
                    ("status",           "VARCHAR(20)",  "NO",  "draft | pending | approved | rejected | returned | po_issued"),
                    ("created_at",       "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",       "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "procurement_items",
                "Line items on a procurement request",
                [
                    ("id",                     "SERIAL",       "NO",  "PK"),
                    ("procurement_request_id", "INT",          "NO",  "FK → procurement_requests.id"),
                    ("description",            "VARCHAR(255)", "NO",  ""),
                    ("quantity",               "INT",          "NO",  ""),
                    ("unit_price",             "NUMERIC(15,2)","YES", ""),
                    ("total_price",            "NUMERIC(15,2)","YES", "computed: qty × unit_price"),
                ],
            ),
            (
                "purchase_orders",
                "PO issued after procurement request is fully approved",
                [
                    ("id",                     "SERIAL",       "NO",  "PK"),
                    ("po_number",              "VARCHAR(20)",  "NO",  "UNIQUE  e.g. PO-2026-001"),
                    ("procurement_request_id", "INT",          "NO",  "FK → procurement_requests.id"),
                    ("vendor_id",              "INT",          "NO",  "FK → vendors.id  (permanent or temporary)"),
                    ("total_amount",           "NUMERIC(15,2)","NO",  ""),
                    ("currency",               "VARCHAR(5)",   "NO",  "default NGN"),
                    ("issued_by",              "INT",          "NO",  "FK → employees.id"),
                    ("issued_at",              "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("status",                 "VARCHAR(20)",  "NO",  "issued | delivered | cancelled"),
                    ("notes",                  "TEXT",         "YES", ""),
                ],
            ),
        ],
    },

    # ── 6. ASSETS ─────────────────────────────────────────────────────────────
    {
        "sheet_name":    "Assets",
        "display_title": "Asset Management",
        "accent":        C["amber"],
        "hidden":        False,
        "tables": [
            (
                "asset_categories",
                "Asset type classification",
                [
                    ("id",          "SERIAL",       "NO",  "PK"),
                    ("name",        "VARCHAR(100)", "NO",  ""),
                    ("description", "TEXT",         "YES", ""),
                    ("created_by",  "INT",          "NO",  "FK → employees.id"),
                    ("created_at",  "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "assets",
                "Individual asset register",
                [
                    ("id",             "SERIAL",       "NO",  "PK"),
                    ("asset_no",       "VARCHAR(20)",  "NO",  "UNIQUE  e.g. AST-001"),
                    ("category_id",    "INT",          "NO",  "FK → asset_categories.id"),
                    ("name",           "VARCHAR(200)", "NO",  ""),
                    ("description",    "TEXT",         "YES", ""),
                    ("serial_number",  "VARCHAR(100)", "YES", ""),
                    ("purchase_date",  "DATE",         "YES", ""),
                    ("purchase_price", "NUMERIC(15,2)","YES", ""),
                    ("current_value",  "NUMERIC(15,2)","YES", ""),
                    ("status",         "VARCHAR(20)",  "NO",  "available | assigned | under_repair | disposed"),
                    ("location",       "VARCHAR(150)", "YES", ""),
                    ("added_by",       "INT",          "NO",  "FK → employees.id"),
                    ("created_at",     "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",     "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "asset_requests",
                "Employee request to purchase, borrow, or repair an asset",
                [
                    ("id",           "SERIAL",       "NO",  "PK"),
                    ("reference",    "VARCHAR(20)",  "NO",  "UNIQUE  e.g. AR-2026-001  — human-readable identifier shown to users"),
                    ("requester_id", "INT",          "NO",  "FK → employees.id"),
                    ("asset_id",     "INT",          "YES", "FK → assets.id  (null if requesting new purchase)"),
                    ("request_type", "VARCHAR(20)",  "NO",  "purchase | loan | repair"),
                    ("description",  "TEXT",         "YES", ""),
                    ("status",       "VARCHAR(20)",  "NO",  "draft | pending | approved | rejected | returned | allocated"),
                    ("created_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "asset_assignment_logs",
                "Tracks who an asset was assigned to and when it was returned",
                [
                    ("id",           "SERIAL",      "NO",  "PK"),
                    ("asset_id",     "INT",         "NO",  "FK → assets.id"),
                    ("assigned_to",  "INT",         "NO",  "FK → employees.id"),
                    ("assigned_by",  "INT",         "NO",  "FK → employees.id"),
                    ("assigned_at",  "TIMESTAMPTZ", "NO",  ""),
                    ("returned_at",  "TIMESTAMPTZ", "YES", "null until returned"),
                    ("notes",        "TEXT",        "YES", ""),
                ],
            ),
            (
                "asset_maintenance_logs",
                "Maintenance and repair history per asset",
                [
                    ("id",               "SERIAL",       "NO",  "PK"),
                    ("asset_id",         "INT",          "NO",  "FK → assets.id"),
                    ("logged_by",        "INT",          "NO",  "FK → employees.id"),
                    ("description",      "TEXT",         "NO",  ""),
                    ("cost",             "NUMERIC(15,2)","YES", ""),
                    ("maintenance_date", "DATE",         "NO",  ""),
                    ("next_due_date",    "DATE",         "YES", ""),
                ],
            ),
        ],
    },

    # ── 7. INTRANET ───────────────────────────────────────────────────────────
    {
        "sheet_name":    "Intranet",
        "display_title": "Intranet Content",
        "accent":        C["purple"],
        "hidden":        False,
        "tables": [
            (
                "intranet_announcements",
                "News and announcements published on the intranet",
                [
                    ("id",           "SERIAL",       "NO",  "PK"),
                    ("title",        "VARCHAR(255)", "NO",  ""),
                    ("body",         "TEXT",         "NO",  ""),
                    ("category",     "VARCHAR(60)",  "NO",  "Company News | Announcement | Policy Update | Safety | Events | Project Update"),
                    ("cover_image",  "INT",          "YES", "FK → documents.id"),
                    ("is_published", "BOOLEAN",      "NO",  "default false"),
                    ("published_at", "TIMESTAMPTZ",  "YES", ""),
                    ("created_by",   "INT",          "NO",  "FK → employees.id"),
                    ("created_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "intranet_events",
                "Company events listed on the intranet events page",
                [
                    ("id",           "SERIAL",       "NO",  "PK"),
                    ("title",        "VARCHAR(255)", "NO",  ""),
                    ("description",  "TEXT",         "YES", ""),
                    ("event_type",   "VARCHAR(40)",  "NO",  "Town Hall | Training | Deadline | Workshop | Social"),
                    ("location",     "VARCHAR(200)", "YES", ""),
                    ("event_date",   "DATE",         "NO",  ""),
                    ("color",        "VARCHAR(10)",  "YES", "hex color for UI card"),
                    ("is_published", "BOOLEAN",      "NO",  "default false"),
                    ("created_by",   "INT",          "NO",  "FK → employees.id"),
                    ("created_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "intranet_spotlight",
                "Employee spotlight and recognition — covers employee of the month, long service, achievements",
                [
                    ("id",           "SERIAL",       "NO",  "PK"),
                    ("employee_id",  "INT",          "NO",  "FK → employees.id  (who is spotlighted)"),
                    ("title",        "VARCHAR(150)", "NO",  "e.g. Employee of the Month — June 2026"),
                    ("message",      "TEXT",         "NO",  ""),
                    ("photo_id",     "INT",          "YES", "FK → documents.id  (spotlight photo)"),
                    ("category",     "VARCHAR(40)",  "NO",  "employee_of_month | long_service | new_joiner | achievement"),
                    ("month",        "INT",          "YES", "1–12  (for employee_of_month)"),
                    ("year",         "INT",          "YES", "e.g. 2026"),
                    ("is_published", "BOOLEAN",      "NO",  "default false"),
                    ("published_at", "TIMESTAMPTZ",  "YES", ""),
                    ("created_by",   "INT",          "NO",  "FK → employees.id"),
                    ("created_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "intranet_message_from",
                "Message from MD or leadership — typically one active record at a time",
                [
                    ("id",           "SERIAL",       "NO",  "PK"),
                    ("author_id",    "INT",          "NO",  "FK → employees.id  (e.g. MD or Dept Head)"),
                    ("title",        "VARCHAR(200)", "NO",  "e.g. Message from the MD — June 2026"),
                    ("body",         "TEXT",         "NO",  ""),
                    ("photo_id",     "INT",          "YES", "FK → documents.id  (author headshot)"),
                    ("is_published", "BOOLEAN",      "NO",  "default false"),
                    ("published_at", "TIMESTAMPTZ",  "YES", ""),
                    ("created_by",   "INT",          "NO",  "FK → employees.id"),
                    ("created_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "intranet_quick_links",
                "Quick-access links shown on the intranet home page",
                [
                    ("id",          "SERIAL",       "NO",  "PK"),
                    ("label",       "VARCHAR(100)", "NO",  "e.g. Submit Expense, Book a Meeting Room"),
                    ("url",         "TEXT",         "NO",  ""),
                    ("icon",        "VARCHAR(60)",  "YES", "icon name for UI"),
                    ("category",    "VARCHAR(80)",  "YES", "HR | Finance | IT | General"),
                    ("order_index", "INT",          "NO",  "controls display order"),
                    ("is_active",   "BOOLEAN",      "NO",  "default true"),
                    ("created_by",  "INT",          "NO",  "FK → employees.id"),
                    ("created_at",  "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "intranet_feedback",
                "Staff feedback and suggestions submitted via the intranet",
                [
                    ("id",           "SERIAL",       "NO",  "PK"),
                    ("submitted_by", "INT",          "YES", "FK → employees.id  (null if anonymous)"),
                    ("category",     "VARCHAR(80)",  "NO",  "General | IT | HR | Suggestion | Complaint"),
                    ("subject",      "VARCHAR(200)", "NO",  ""),
                    ("message",      "TEXT",         "NO",  ""),
                    ("is_anonymous", "BOOLEAN",      "NO",  "default false"),
                    ("status",       "VARCHAR(20)",  "NO",  "open | in_review | resolved | closed"),
                    ("resolved_by",  "INT",          "YES", "FK → employees.id"),
                    ("resolved_at",  "TIMESTAMPTZ",  "YES", ""),
                    ("created_at",   "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
            (
                "intranet_faq",
                "FAQ entries managed via admin portal",
                [
                    ("id",          "SERIAL",       "NO",  "PK"),
                    ("question",    "TEXT",         "NO",  ""),
                    ("answer",      "TEXT",         "NO",  ""),
                    ("category",    "VARCHAR(80)",  "NO",  "IT Support | HR & Payroll | HSE | Procurement | General"),
                    ("order_index", "INT",          "NO",  "controls display order within category"),
                    ("is_published","BOOLEAN",      "NO",  "default false"),
                    ("created_by",  "INT",          "NO",  "FK → employees.id"),
                    ("created_at",  "TIMESTAMPTZ",  "NO",  "default now()"),
                    ("updated_at",  "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
        ],
    },

    # ── 8. NOTIFICATIONS ─────────────────────────────────────────────────────
    {
        "sheet_name":    "Notifications",
        "display_title": "Notifications",
        "accent":        C["teal"],
        "hidden":        False,
        "tables": [
            (
                "notifications",
                "In-app notifications sent to employees — also used to trigger emails",
                [
                    ("id",             "SERIAL",       "NO",  "PK"),
                    ("recipient_id",   "INT",          "NO",  "FK → employees.id"),
                    ("type",           "VARCHAR(40)",  "NO",  "approval_required | approved | rejected | returned | info"),
                    ("title",          "VARCHAR(200)", "NO",  ""),
                    ("message",        "TEXT",         "NO",  ""),
                    ("reference_type", "VARCHAR(50)",  "YES", "procurement | asset"),
                    ("reference_id",   "INT",          "YES", "PK of the originating record"),
                    ("is_read",        "BOOLEAN",      "NO",  "default false"),
                    ("created_at",     "TIMESTAMPTZ",  "NO",  "default now()"),
                ],
            ),
        ],
    },

]

# ── Sheet builders ────────────────────────────────────────────────────────────

def make_index_sheet(wb, groups):
    ws = wb.active
    ws.title = "Index"

    # Banner
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "PORTLAND GAS LIMITED — DATABASE DESIGN"
    c.fill = fill(C["dark"])
    c.font = Font(bold=True, color=C["gold"], size=15, name="Calibri")
    c.alignment = center()
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "Schema Reference  ·  v8.0  ·  June 2026"
    c.fill = fill(C["purple"])
    c.font = font(color=C["white"], size=10)
    c.alignment = center()
    ws.row_dimensions[2].height = 20

    # FK Rule callout
    ws.merge_cells("A4:F4")
    c = ws["A4"]
    c.value = "⚡  FK RULE:  employees.user_id → users.id  is the ONLY link to users.  ALL other FKs point to employees.id"
    c.fill = fill(C["gold_lt"])
    c.font = Font(bold=True, color=C["amber"], size=10, name="Calibri")
    c.alignment = center()
    ws.row_dimensions[4].height = 22

    # Table of contents header
    headers = ["Module", "Sheet", "Table", "Rows (est.)", "Purpose"]
    col_widths = [20, 16, 28, 14, 60]
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=6, column=i, value=h)
        c.fill = fill(C["header_row"])
        c.font = font(bold=True, color=C["white"])
        c.alignment = center()
        c.border = border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[6].height = 20

    row = 7
    for g in groups:
        if g is None:
            continue
        # Skip hidden groups from the index
        if g.get("hidden", False):
            continue
        sheet_name    = g["sheet_name"]
        display_title = g["display_title"]
        tables        = g["tables"]
        for table_name, desc, cols in tables:
            est_rows = {
                "users": "~200", "employees": "~200",
                "documents": "thousands", "approval_history": "thousands",
                "workflow_audit_trail": "thousands",
                "approval_step_assignments": "thousands",
                "notifications": "thousands",
            }.get(table_name, "hundreds")

            vals = [display_title, sheet_name, table_name, est_rows, desc]
            for i, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=i, value=v)
                c.fill = fill("F8F4FF") if row % 2 == 0 else fill(C["white"])
                c.font = font(bold=(i == 3), color=C["dark"] if i != 3 else C["purple"])
                c.alignment = left() if i != 4 else center()
                c.border = border()
            ws.row_dimensions[row].height = 18
            row += 1

    # Legend
    row += 1
    legends = [
        ("PK",   C["pk"],    "Primary Key"),
        ("FK",   C["fk"],    "Foreign Key — references another table"),
        ("NULL", "FEF9C3",   "Column is nullable (YES in Nullable column)"),
    ]
    ws.cell(row=row, column=1, value="Legend").font = font(bold=True)
    row += 1
    for lbl, color, desc in legends:
        c = ws.cell(row=row, column=1, value=f"  {lbl}")
        c.fill = fill(color)
        c.font = font(bold=True, color=C["dark"])
        c.border = border()
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.font = font(color=C["slate"])
        c2.border = border()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False


def make_group_sheet(wb, sheet_name, display_title, accent, tables):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"PORTLAND GAS — {display_title.upper()}"
    c.fill = fill(C["dark"])
    c.font = Font(bold=True, color=C["gold"], size=13, name="Calibri")
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    col_widths = [28, 18, 10, 40]
    col_headers = ["Column", "Type", "Nullable", "Notes / FK"]
    for i, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    current_row = 3

    for table_name, desc, cols in tables:
        # Table name banner
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        c = ws.cell(row=current_row, column=1, value=f"  {table_name}")
        c.fill = fill(accent)
        c.font = Font(bold=True, color=C["white"], size=11, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Description
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        c = ws.cell(row=current_row, column=1, value=f"  {desc}")
        c.fill = fill("F1F5F9")
        c.font = Font(italic=True, color=C["slate"], size=9, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Column headers
        for i, h in enumerate(col_headers, start=1):
            c = ws.cell(row=current_row, column=i, value=h)
            c.fill = fill(C["col_header"])
            c.font = font(bold=True, color=C["white"], size=9)
            c.alignment = center()
            c.border = border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Rows
        for r_idx, (col_name, col_type, nullable, notes) in enumerate(cols):
            is_pk = notes.startswith("PK")
            is_fk = "FK" in notes

            row_fill = C["gold_lt"] if is_pk else ("F5F0FF" if is_fk else (C["white"] if r_idx % 2 == 0 else C["slate_lt"]))

            vals = [col_name, col_type, nullable, notes]
            for i, v in enumerate(vals, start=1):
                c = ws.cell(row=current_row, column=i, value=v)
                c.fill = fill(row_fill)
                if i == 1 and is_pk:
                    c.font = Font(bold=True, color="92400E", size=10, name="Calibri")
                elif i == 1 and is_fk:
                    c.font = Font(bold=True, color=C["blue"], size=10, name="Calibri")
                elif i == 3 and nullable == "YES":
                    c.font = font(color=C["null_col"], italic=True)
                elif i == 4 and is_fk:
                    c.font = font(color=C["blue"], italic=True, size=9)
                else:
                    c.font = font(size=10, color=C["dark"] if i < 3 else C["slate"])
                c.alignment = left() if i != 3 else center()
                c.border = border()
            ws.row_dimensions[current_row].height = 17
            current_row += 1

        current_row += 2  # spacer between tables

    ws.freeze_panes = "A4"
    return ws


def make_future_sheet(wb):
    ws = wb.create_sheet(title="Future Extras")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 55

    # Banner
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "PORTLAND GAS — FUTURE EXTRAS (NOT YET IMPLEMENTED)"
    c.fill = fill(C["slate"])
    c.font = Font(bold=True, color=C["gold"], size=13, name="Calibri")
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    # Intro
    ws.merge_cells("A3:D3")
    c = ws["A3"]
    c.value = "  These tables are documented here for future consideration. Do NOT build until the current workflow_steps assignee_type approach becomes unmanageable."
    c.fill = fill("FFFBEA")
    c.font = Font(italic=True, color=C["amber"], size=10, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 36

    # Section title
    ws.merge_cells("A5:D5")
    c = ws["A5"]
    c.value = "  ALTERNATIVE APPROACH: resolution_method + resolution_key on workflow_steps"
    c.fill = fill(C["blue"])
    c.font = Font(bold=True, color=C["white"], size=11, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = 24

    def section_header(ws, row, text, color):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f"  {text}")
        c.fill = fill(color)
        c.font = Font(bold=True, color=C["white"], size=11, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24
        return row + 1

    def write_rows(ws, row, rows):
        for label, text in rows:
            c1 = ws.cell(row=row, column=1, value=f"  {label}" if label else "")
            c1.font = Font(bold=bool(label), color=C["dark"], size=10, name="Calibri")
            c1.fill = fill(C["white"] if row % 2 == 0 else C["slate_lt"])
            c1.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            c2 = ws.cell(row=row, column=2, value=text)
            c2.font = Font(color=C["slate"], size=10, name="Calibri")
            c2.fill = fill(C["white"] if row % 2 == 0 else C["slate_lt"])
            c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 22
            row += 1
        return row + 1

    row = 6

    # ── Section 1: Expression approach ───────────────────────────────────────
    row = section_header(ws, row, "APPROACH 1 — Expression-based assignee on workflow_steps  (industry standard)", C["blue"])

    row = write_rows(ws, row, [
        ("WHAT IT IS", "Used by ServiceNow, Camunda, Jira Service Management, Nintex, and most enterprise workflow engines. Instead of adding a new assignee_type value for every org-chart relationship, you replace the multiple columns (role, employee_id, group_id) with just two: assignee_type and assignee_value."),
        ("", ""),
        ("COLUMN CHANGE", "Remove: role, employee_id, group_id from workflow_steps"),
        ("", "Add:    assignee_type  VARCHAR(20)   —   role | user | group | expression"),
        ("", "        assignee_value VARCHAR(200)  —   the value depending on type"),
        ("", ""),
        ("assignee_type=role",       "assignee_value = finance_manager  →  engine queries employees WHERE role = 'finance_manager'"),
        ("assignee_type=user",       "assignee_value = 5                →  engine uses employee ID 5 directly"),
        ("assignee_type=group",      "assignee_value = 3                →  engine shows picker filtered to approver_group ID 3"),
        ("assignee_type=expression", "assignee_value = requester.line_manager           →  engine reads requester.line_manager_id"),
        ("",                         "assignee_value = requester.department.hod         →  engine finds HOD of requester's department"),
        ("",                         "assignee_value = requester.line_manager.line_manager  →  skip-level manager"),
        ("", ""),
        ("ADMIN UX", "Admins never type expressions manually. The frontend shows a friendly dropdown:"),
        ("",         "  'Org Chart Relationship' → sub-dropdown: Line Manager | Head of Department | Skip-Level"),
        ("",         "Each selection maps to an expression string behind the scenes. Admin sees plain English only."),
        ("", ""),
        ("BENEFIT",       "Adding a new org-chart relationship = add one dropdown option on the frontend + one line in the engine evaluator. No schema change, no migration needed."),
        ("WHEN TO SWITCH", "When assignee_type values exceed 6 or the workflow_steps table starts feeling cluttered with nulls across role/employee_id/group_id columns."),
    ])

    # ── Section 2: Original resolution_method approach ────────────────────────
    row = section_header(ws, row, "APPROACH 2 — resolution_method + resolution_key  (earlier iteration, superseded by Approach 1)", C["slate"])

    row = write_rows(ws, row, [
        ("NOTE",                    "This was an earlier iteration of the same idea. Approach 1 (expression) is cleaner and more widely used. Keep this only as historical context."),
        ("", ""),
        ("resolution_method values", "org_chart  |  role_lookup  |  fixed  |  picker"),
        ("resolution_key examples",  "line_manager, head_of_department  (for org_chart)"),
        ("",                         "finance_manager, md  (for role_lookup)"),
        ("",                         "null — employee_id column used  (for fixed)"),
        ("",                         "null — group_id column used  (for picker)"),
    ])

    return ws


# ── Build workbook ────────────────────────────────────────────────────────────

make_index_sheet(wb, GROUPS)

for g in GROUPS:
    if g is None:
        continue
    ws = make_group_sheet(wb, g["sheet_name"], g["display_title"], g["accent"], g["tables"])
    if g.get("hidden", False):
        ws.sheet_state = "hidden"

future_ws = make_future_sheet(wb)
future_ws.sheet_state = "hidden"

out = "/Users/nubiaville/Documents/portland-gas-ops/Portland_Gas_DB_Design.xlsx"
wb.save(out)

total_tables = sum(len(g["tables"]) for g in GROUPS if g is not None)
total_cols   = sum(len(t[2]) for g in GROUPS if g is not None for t in g["tables"])
print(f"Done. {total_tables} tables · {total_cols} columns → {out}")
